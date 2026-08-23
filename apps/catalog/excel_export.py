"""FEAT-078 — export Excel du catalogue (une ligne par exemplaire).

Le fichier produit porte exactement les colonnes reconnues par l'import
(FEAT-050/053) et par la mise à jour d'exemplaires (FEAT-079), plus les codes
qui identifient l'exemplaire. Il est donc réutilisable tel quel :

    exporter → corriger dans Excel → « Mettre à jour les exemplaires »

Une ligne = un exemplaire, pas une notice : l'emplacement, l'état, la
provenance et le code externe appartiennent à l'exemplaire. Une notice à trois
exemplaires sort donc sur trois lignes, dont les colonnes de fiche (titre,
auteur…) sont identiques.

L'export est **synchrone** (téléchargement direct), contrairement à la
vérification et à l'import : il ne fait aucun appel réseau, seulement une
lecture de base, et une bibliothèque Ofelia se compte en milliers
d'exemplaires, pas en millions.
"""
from __future__ import annotations

import io

# Ordre des colonnes du fichier exporté. Les deux premières identifient
# l'exemplaire pour la mise à jour ; les suivantes sont exactement celles que
# l'import et la mise à jour savent lire.
EXPORT_COLUMNS = [
    "OFELIA_CODE",
    "INTERNAL_ID",
    "EXTERNAL_CODE",
    "ISBN",
    "TITLE",
    "AUTHOR",
    "CATEGORY",
    "CATEGORY_ABBR",
    "TYPE",
    "EDITOR",
    "YEAR",
    "LANGUAGE",
    "TAGS",
    "CONDITION",
    "PROVENANCE",
    "LOCATION",
]

# Largeurs de colonne (en caractères), dans l'ordre de EXPORT_COLUMNS. Sans
# elles le fichier s'ouvre sur seize colonnes de 8 caractères où rien n'est
# lisible — et le premier geste du bibliothécaire est de tout élargir à la main.
_COLUMN_WIDTHS = [16, 18, 16, 15, 44, 30, 24, 14, 16, 24, 7, 10, 24, 10, 14, 12]


def items_queryset():
    """Tous les exemplaires, prêts à être exportés (1 requête + prefetch)."""
    from .models import Item

    return (
        Item.objects.select_related(
            "record", "record__category", "location", "provenance"
        )
        .prefetch_related("record__authors", "record__tags")
        .order_by("record__title", "internal_id")
    )


def export_row(item) -> list:
    """Une ligne d'export pour `item`, dans l'ordre de EXPORT_COLUMNS.

    TYPE et CONDITION sortent avec leur libellé traduit dans la langue de
    l'utilisateur (« Livre », « Libro », « Boky ») : c'est le fichier d'un
    bibliothécaire, pas un export machine. La relecture côté import/mise à jour
    accepte les libellés de toutes les langues de l'instance
    (`excel_catalog._resolve_document_type`), donc l'aller-retour tient quelle
    que soit la langue d'export.
    """
    record = item.record
    return [
        item.ean13,
        item.internal_id,
        item.external_code,
        record.isbn_13 or record.isbn_10 or "",
        record.title,
        "; ".join(a.full_name for a in record.authors.all()),
        record.category.name if record.category else "",
        record.category.abbreviation if record.category else "",
        record.get_document_type_display(),
        record.publisher,
        record.publication_year,
        record.language,
        ", ".join(t.name for t in record.tags.all()),
        item.get_state_display(),
        item.provenance.code if item.provenance else "",
        item.location.code if item.location else "",
    ]


def build_catalog_workbook() -> bytes:
    """Construit le .xlsx de tout le catalogue et renvoie ses octets.

    Mode `write_only` d'openpyxl : les lignes partent au fichier au fil de
    l'itération au lieu de s'empiler en mémoire — sur une Box à 4 Go, un
    catalogue de 10 000 exemplaires n'a pas à tenir deux fois en RAM.
    """
    import openpyxl
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="Catalogue")
    bold = Font(bold=True)

    for idx, width in enumerate(_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    # Les en-têtes restent visibles quand on descend dans un fichier long.
    ws.freeze_panes = "A2"

    header = []
    for name in EXPORT_COLUMNS:
        cell = WriteOnlyCell(ws, value=name)
        cell.font = bold
        header.append(cell)
    ws.append(header)

    for item in items_queryset().iterator(chunk_size=500):
        ws.append(export_row(item))

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()
