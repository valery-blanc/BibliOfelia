"""Service de catalogage Excel (FEAT-050).

Deux modes, tous deux exécutés en tâche django-q2 via ``run_excel_catalog_job`` :

- **VERIFY** : annote un fichier Excel avec ce que les sources en ligne
  connaissent du livre — d'abord par ISBN (passe 1), puis par titre + auteur en
  repli (passe 2, réordonnée par score fuzzy local). Produit un fichier
  téléchargeable. Aucun effet de bord sur le catalogue.
- **IMPORT** : matérialise une liste d'ISBN en notices + exemplaires via une
  ``ScanSession`` virtuelle, puis ``finalize_scan_session`` (réutilise le
  pipeline FEAT-021 / FEAT-046).
- **UPDATE** (FEAT-079) : met à jour des exemplaires **existants**, retrouvés
  par leur code Ofelia et/ou leur code externe. Ne crée jamais rien : une ligne
  dont l'exemplaire est introuvable est signalée et laissée de côté.

La validation du fichier (``validate_xlsx``) est faite côté vue, avant la
création du job ; le runner ré-ouvre ensuite ``job.uploaded_file``.
"""
from __future__ import annotations

import io
import logging
import unicodedata
from concurrent.futures import ThreadPoolExecutor

from django.core.files.base import ContentFile
from django.utils import timezone

from .enrichment import _try_sources
from .models import (
    Category,
    DocumentType,
    ExcelCatalogJob,
    ExcelJobMode,
    ExcelJobState,
    ItemState,
    ScanItem,
    ScanKind,
    ScanSession,
    ScanSessionState,
    Tag,
)
from .lookup import is_valid_external_code, normalize_external_code
from .openlibrary import normalize_isbn
from .sources import SEARCHES, SOURCE_LABELS
from .sources._fuzzy import CONFIDENCE_FLOOR, HIGHLIGHT_BELOW, best_candidate

logger = logging.getLogger(__name__)

# Garde-fous d'upload.
MAX_FILE_BYTES = 5 * 1024 * 1024
MAX_ROWS = 10_000

# Les 4 sources interrogées, dans l'ordre de préférence (1re non vide gagne).
_SOURCE_ORDER = ["openlibrary", "google_books", "bnf", "bne"]

# Colonnes ajoutées en queue par le mode VERIFY (toujours dans cet ordre).
VERIFY_OUTPUT_COLUMNS = [
    "TITLE_FOUND_BY_ISBN",
    "AUTHOR_FOUND_BY_ISBN",
    "SOURCE_BY_ISBN",
    "ISBN_FOUND_BY_TA",
    "TITLE_FOUND_BY_TA",
    "AUTHOR_FOUND_BY_TA",
    "SOURCE_BY_TA",
    "CONFIDENCE",
]


# FEAT-053 : colonnes optionnelles d'affectation de la fiche catalogue en mode
# IMPORT. Une colonne présente avec une cellule non vide écrase le champ de la
# notice (même existante) ; une cellule vide laisse l'existant intact. AUTHOR et
# TAGS remplacent (pas de fusion). Toutes optionnelles.
IMPORT_OVERRIDE_COLUMNS = [
    "title",
    "author",
    "category",
    "type",
    "editor",
    "year",
    "language",
    "tags",
    "condition",  # → état de l'exemplaire (Item.state)
    "external_code",  # FEAT-063 → code Ofelia externe de l'exemplaire
    "provenance",     # FEAT-064 → provenance de l'exemplaire
    "category_abbr",  # FEAT-067 → abréviation (cote) de la catégorie
]

# FEAT-079 : colonnes qui identifient l'exemplaire à mettre à jour. Au moins
# l'une d'elles doit être présente dans le fichier ; le code Ofelia l'emporte
# sur le code externe quand les deux sont renseignés sur une ligne.
UPDATE_KEY_COLUMNS = ["ofelia_code", "internal_id", "external_code"]

# FEAT-079 : colonnes modifiables en mise à jour. Ce sont celles de l'import,
# plus LOCATION et ISBN — en import ces deux-là ne sont pas des « overrides »
# (ISBN est la clé, LOCATION est posée à la création de l'exemplaire), en mise
# à jour ce sont des champs comme les autres.
UPDATE_OVERRIDE_COLUMNS = IMPORT_OVERRIDE_COLUMNS + ["location", "isbn"]

# En-têtes acceptés pour les colonnes dont le nom « naturel » varie d'un
# fichier à l'autre. Clés déjà normalisées par `_norm` (minuscules, sans
# accents). Le nom canonique reste toujours accepté.
_COLUMN_ALIASES = {
    "external_code": [
        "code_externe",
        "code externe",
        "code_ofelia_externe",
        "code ofelia externe",
        "ofelia_ext",
        "externalcode",
    ],
    "provenance": ["origine"],
    # FEAT-079 : dans l'UI, « code Ofelia » désigne l'EAN13 290… et « code
    # interne » l'identifiant OFL-… ; les deux sont acceptés comme clé, et
    # `_find_item_by_ofelia_code` essaie l'un puis l'autre quelle que soit la
    # colonne d'où vient la valeur.
    "ofelia_code": [
        "code_ofelia",
        "code ofelia",
        "ofelia",
        "ean13",
        "ean_13",
    ],
    "internal_id": [
        "code_interne",
        "code interne",
        "internalid",
        "id_ofelia",
    ],
    "category_abbr": [
        "abbreviation",
        "abreviation",
        "categorie_abregee",
        "categorie abregee",
        "category_abbreviation",
        "cat_abbr",
    ],
}


def _resolve_column(headers: dict[str, int], name: str) -> int | None:
    """Index de la colonne `name`, en acceptant ses alias d'en-tête."""
    idx = headers.get(name)
    if idx:
        return idx
    for alias in _COLUMN_ALIASES.get(name, ()):
        idx = headers.get(alias)
        if idx:
            return idx
    return None

# Garde-fous tags (alignés sur enrichment.py).
_MAX_TAGS_PER_RECORD = 10
_MAX_TAG_LENGTH = 40

# Résolution TYPE (colonne Excel) → DocumentType. Clés normalisées via _norm
# (minuscules, sans accents) : accepte le code interne ou un libellé FR courant.
_DOCUMENT_TYPE_ALIASES = {
    "book": DocumentType.BOOK,
    "livre": DocumentType.BOOK,
    "magazine_issue": DocumentType.MAGAZINE_ISSUE,
    "magazine": DocumentType.MAGAZINE_ISSUE,
    "numero de magazine": DocumentType.MAGAZINE_ISSUE,
    "revue": DocumentType.MAGAZINE_ISSUE,
    "periodique": DocumentType.MAGAZINE_ISSUE,
    "newspaper": DocumentType.NEWSPAPER,
    "journal": DocumentType.NEWSPAPER,
    "comic": DocumentType.COMIC,
    "bd": DocumentType.COMIC,
    "manga": DocumentType.COMIC,
    "bd / manga": DocumentType.COMIC,
    "bd/manga": DocumentType.COMIC,
    "bande dessinee": DocumentType.COMIC,
    "audio_cd": DocumentType.AUDIO_CD,
    "cd": DocumentType.AUDIO_CD,
    "cd audio": DocumentType.AUDIO_CD,
    "audio": DocumentType.AUDIO_CD,
    "other": DocumentType.OTHER,
    "autre": DocumentType.OTHER,
}

# Résolution CONDITION (colonne Excel) → ItemState (état exemplaire).
_ITEM_STATE_ALIASES = {
    "new": ItemState.NEW,
    "neuf": ItemState.NEW,
    "good": ItemState.GOOD,
    "bon": ItemState.GOOD,
    "worn": ItemState.WORN,
    "use": ItemState.WORN,  # « usé » normalisé
    "damaged": ItemState.DAMAGED,
    "abime": ItemState.DAMAGED,  # « abîmé » normalisé
}


def _norm(value: str) -> str:
    """Normalise un nom de colonne : minuscules, sans accents, sans espaces."""
    s = unicodedata.normalize("NFKD", str(value or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip().lower()


# FEAT-078 : libellés traduits → valeur, construits une fois par processus.
_LABEL_ALIAS_CACHE: dict[str, dict[str, str]] = {}


def _translated_label_aliases(key: str, choices) -> dict[str, str]:
    """Libellés de `choices` normalisés → valeur, dans toutes les langues.

    L'export (FEAT-078) écrit TYPE et CONDITION avec leur libellé traduit dans
    la langue du bibliothécaire. Sans cette table, un fichier exporté en
    espagnol reviendrait en mise à jour avec « Libro » / « Bueno » partout,
    donc TYPE_UNKNOWN et CONDITION_UNKNOWN sur chaque ligne — l'aller-retour ne
    marcherait qu'en français.
    """
    cached = _LABEL_ALIAS_CACHE.get(key)
    if cached is not None:
        return cached
    from django.conf import settings
    from django.utils.translation import override

    mapping: dict[str, str] = {}
    for lang_code, _lang_label in settings.LANGUAGES:
        with override(lang_code):
            for value, label in choices:
                norm = _norm(str(label))
                if norm:
                    mapping.setdefault(norm, value)
    _LABEL_ALIAS_CACHE[key] = mapping
    return mapping


def _translated_name_fields() -> list[str]:
    """Champs `name_<lang>` créés par modeltranslation, pour Category et Tag."""
    from django.conf import settings

    return [f"name_{code.replace('-', '_')}" for code, _label in settings.LANGUAGES]


def _resolve_category(value: str):
    """Catégorie désignée par son nom (dans n'importe quelle langue) ou son code.

    `Category.name` est un champ traduit (modeltranslation) : l'export FEAT-078
    écrit celui de la langue du bibliothécaire, alors que le job, lui, tourne
    dans le worker django-q2, en français. Une recherche sur `name` seul
    renverrait donc CATEGORY_UNKNOWN sur chaque ligne d'un fichier exporté en
    espagnol. Le nom passe avant le code : c'est ce que le bibliothécaire tape.
    """
    from django.db.models import Q

    value = (value or "").strip()
    if not value:
        return None
    query = Q()
    for field in _translated_name_fields():
        query |= Q(**{f"{field}__iexact": value})
    return (
        Category.objects.filter(query).first()
        or Category.objects.filter(code__iexact=value).first()
    )


def _get_or_create_tag(name: str):
    """Tag portant `name` dans une langue quelconque, créé à défaut.

    Même raison que `_resolve_category` : sans la recherche multi-langue, un
    fichier exporté en espagnol recréerait chaque tag en double, le libellé
    espagnol atterrissant dans le champ français.
    """
    from django.db.models import Q

    query = Q()
    for field in _translated_name_fields():
        query |= Q(**{f"{field}__iexact": name})
    existing = Tag.objects.filter(query).first()
    if existing is not None:
        return existing
    return Tag.objects.create(name=name)


def _resolve_document_type(value: str) -> str | None:
    """FEAT-053 : mappe une valeur TYPE (code ou libellé) → DocumentType.

    Les alias FR écrits à la main l'emportent (ils couvrent les variantes de
    saisie : « BD », « manga », « périodique »…) ; à défaut on cherche parmi
    les libellés officiels de toutes les langues (FEAT-078). None si inconnue.
    """
    norm = _norm(value)
    return _DOCUMENT_TYPE_ALIASES.get(norm) or _translated_label_aliases(
        "document_type", DocumentType.choices
    ).get(norm)


def _resolve_item_state(value: str) -> str | None:
    """FEAT-053 : mappe une valeur CONDITION (code ou libellé) → ItemState.
    None si non reconnue. Voir `_resolve_document_type` pour les langues."""
    norm = _norm(value)
    return _ITEM_STATE_ALIASES.get(norm) or _translated_label_aliases(
        "item_state", ItemState.choices
    ).get(norm)


def _split_multi(value: str, sep: str) -> list[str]:
    """Découpe une cellule multi-valeurs (auteurs, tags), sans doublon ni vide."""
    seen: list[str] = []
    for part in value.split(sep):
        clean = part.strip()
        if clean and clean not in seen:
            seen.append(clean)
    return seen


def _cell_str(value) -> str:
    """Valeur de cellule en str propre. Gère les ISBN stockés en nombre."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_map(ws) -> dict[str, int]:
    """Mappe nom de colonne normalisé → index de colonne (1-based)."""
    mapping: dict[str, int] = {}
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for idx, value in enumerate(first_row, start=1):
        key = _norm(value)
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def _row_label(row, headers: dict[str, int]) -> str:
    """« Auteur — Titre » d'une ligne, pour identifier une ligne signalée.

    Sert au rapport d'import (BUG-025) : une ligne sans ISBN n'a pas d'autre
    identifiant lisible que son auteur et son titre.
    """
    parts = []
    for name in ("author", "title"):
        col = headers.get(name)
        if col and col <= len(row):
            value = _cell_str(row[col - 1])
            if value:
                parts.append(value)
    return " — ".join(parts)


def required_columns(mode: str) -> list[str]:
    if mode == ExcelJobMode.IMPORT:
        return ["isbn"]
    if mode == ExcelJobMode.UPDATE:
        # FEAT-079 : pas une liste de colonnes toutes obligatoires mais un
        # « au moins une de celles-ci » — traité à part dans `validate_xlsx`.
        return UPDATE_KEY_COLUMNS
    return ["id", "title", "author", "isbn"]


def validate_xlsx(uploaded_file, mode: str) -> list[str]:
    """Valide un fichier uploadé (vue). Retourne la liste des erreurs (vide = OK).

    Vérifie : extension/type, taille, ouverture openpyxl, nombre de lignes,
    présence des colonnes obligatoires.
    """
    import openpyxl
    from django.utils.translation import gettext

    errors: list[str] = []
    name = (getattr(uploaded_file, "name", "") or "").lower()
    if not name.endswith(".xlsx"):
        errors.append(
            "Format non supporté : seul le format .xlsx est accepté en v1 "
            "(.xls, .csv et .ods doivent être convertis au préalable)."
        )
        return errors
    size = getattr(uploaded_file, "size", 0) or 0
    if size > MAX_FILE_BYTES:
        errors.append("Fichier trop volumineux (maximum 5 Mo).")
        return errors
    try:
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception:
        errors.append("Fichier illisible : ce n'est pas un .xlsx valide.")
        return errors
    try:
        ws = wb.active
        if ws is None:
            errors.append("Le classeur ne contient aucune feuille.")
            return errors
        if (ws.max_row or 0) > MAX_ROWS + 1:
            errors.append(f"Trop de lignes (maximum {MAX_ROWS}).")
        headers = _header_map(ws)
        if mode == ExcelJobMode.UPDATE:
            # FEAT-079 : la mise à jour ne crée rien, il lui faut de quoi
            # retrouver l'exemplaire. Une seule des colonnes clés suffit, et
            # les alias d'en-tête sont acceptés (CODE_OFELIA, CODE_EXTERNE…).
            if not any(_resolve_column(headers, c) for c in UPDATE_KEY_COLUMNS):
                errors.append(
                    gettext(
                        "Colonne d'identification manquante : le fichier doit "
                        "porter une colonne OFELIA_CODE (ou INTERNAL_ID) et/ou "
                        "EXTERNAL_CODE, sinon les exemplaires à mettre à jour "
                        "ne peuvent pas être retrouvés."
                    )
                )
        else:
            missing = [c.upper() for c in required_columns(mode) if c not in headers]
            if missing:
                errors.append(
                    "Colonnes obligatoires manquantes : " + ", ".join(missing) + "."
                )
    finally:
        wb.close()
    try:
        uploaded_file.seek(0)
    except Exception:  # pragma: no cover (flux non re-seekable)
        pass
    return errors


def _search_all(title: str, author: str) -> tuple[list[dict], bool]:
    """Interroge les 4 sources `search` en parallèle, agrège les candidats.

    Renvoie ``(candidats, rate_limited)`` — ``rate_limited`` vrai si une source
    a atteint son quota (429) pendant la recherche (BUG-019)."""
    from .sources import SourceRateLimited

    candidates: list[dict] = []
    rate_limited = False

    def _one(name):
        try:
            return SEARCHES[name](title, author, limit=5) or [], False
        except SourceRateLimited:
            logger.info("Source %s search : quota atteint (429) « %s »", name, title)
            return [], True
        except Exception as exc:  # filet : une source qui casse ne bloque pas
            logger.info("Source %s search KO « %s » : %s", name, title, exc)
            return [], False

    with ThreadPoolExecutor(max_workers=len(SEARCHES)) as ex:
        for result, was_limited in ex.map(_one, _SOURCE_ORDER):
            candidates.extend(result)
            rate_limited = rate_limited or was_limited
    return candidates, rate_limited


def _pass1_by_isbn(isbn: str) -> tuple[dict | None, bool]:
    """Passe 1 : interroge les sources par ISBN, renvoie ``(hit, rate_limited)``
    où ``hit`` est la 1re réponse non vide (titre présent) avec sa source, ou
    None. ``rate_limited`` vrai si une source a atteint son quota (429)."""
    responses, rate_limited = _try_sources(isbn, _SOURCE_ORDER, with_rate_limit=True)
    for name in _SOURCE_ORDER:
        data = responses.get(name)
        if data and (data.get("title") or "").strip():
            return {
                "title": (data.get("title") or "").strip(),
                "authors_text": (data.get("authors_text") or "").strip(),
                "source": name,
            }, rate_limited
    return None, rate_limited


def run_verify_job(job: ExcelCatalogJob) -> None:
    """Mode VERIFY : annote l'Excel et écrit ``result_file``."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.load_workbook(job.uploaded_file.path, data_only=True)
    ws = wb.active
    headers = _header_map(ws)
    col_title = headers.get("title")
    col_author = headers.get("author")
    col_isbn = headers.get("isbn")

    # Colonnes de sortie ajoutées en queue.
    start_col = ws.max_column + 1
    bold = Font(bold=True)
    orange = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
    for offset, name in enumerate(VERIFY_OUTPUT_COLUMNS):
        cell = ws.cell(row=1, column=start_col + offset, value=name)
        cell.font = bold
    # En-têtes d'origine en gras aussi.
    for col in range(1, start_col):
        ws.cell(row=1, column=col).font = bold

    out = {n: start_col + i for i, n in enumerate(VERIFY_OUTPUT_COLUMNS)}
    total = ws.max_row - 1 if ws.max_row else 0
    job.total = max(0, total)
    job.save(update_fields=["total"])

    for row in range(2, (ws.max_row or 1) + 1):
        raw_isbn = _cell_str(ws.cell(row=row, column=col_isbn).value) if col_isbn else ""
        title = _cell_str(ws.cell(row=row, column=col_title).value) if col_title else ""
        author = _cell_str(ws.cell(row=row, column=col_author).value) if col_author else ""

        # Ligne vide → on saute sans la compter comme « non trouvée ».
        if not (raw_isbn or title or author):
            continue

        typed_isbn = normalize_isbn(raw_isbn) if raw_isbn else ""
        found_by_isbn = False
        row_rate_limited = False
        if raw_isbn:
            isbn = typed_isbn
            if len(isbn) not in (10, 13):
                ws.cell(row=row, column=out["SOURCE_BY_ISBN"], value="ISBN_INVALID")
            else:
                hit, rl = _pass1_by_isbn(isbn)
                row_rate_limited = row_rate_limited or rl
                if hit:
                    ws.cell(row=row, column=out["TITLE_FOUND_BY_ISBN"], value=hit["title"])
                    ws.cell(row=row, column=out["AUTHOR_FOUND_BY_ISBN"], value=hit["authors_text"])
                    ws.cell(row=row, column=out["SOURCE_BY_ISBN"],
                            value=SOURCE_LABELS.get(hit["source"], hit["source"]))
                    found_by_isbn = True
                    job.matched_by_isbn += 1

        # Passe 2 lancée pour TOUTES les lignes ayant un titre, y compris
        # celles résolues par ISBN : les ISBN sont saisis à la main et peuvent
        # être erronés. La colonne ISBN_FOUND_BY_TA permet alors de comparer
        # l'ISBN trouvé par titre+auteur à celui du fichier (FEAT-050 itér. 2).
        found_by_ta = False
        if title:
            candidates, rl = _search_all(title, author)
            row_rate_limited = row_rate_limited or rl
            best, sc = best_candidate(title, author, candidates)
            if best and sc >= CONFIDENCE_FLOOR:
                isbn_found = best.get("isbn_13") or best.get("isbn_10") or ""
                ta_cell = ws.cell(row=row, column=out["ISBN_FOUND_BY_TA"], value=isbn_found)
                ws.cell(row=row, column=out["TITLE_FOUND_BY_TA"], value=best.get("title", ""))
                ws.cell(row=row, column=out["AUTHOR_FOUND_BY_TA"], value=best.get("authors_text", ""))
                # Source non tracée par candidat → on laisse vide, le score suffit.
                conf_cell = ws.cell(row=row, column=out["CONFIDENCE"], value=sc)
                if sc < HIGHLIGHT_BELOW:
                    conf_cell.fill = orange
                # ISBN saisi ≠ ISBN trouvé par titre+auteur (score fiable) →
                # probable erreur de saisie : on colore la cellule pour la repérer.
                if isbn_found and typed_isbn and isbn_found != typed_isbn and sc >= HIGHLIGHT_BELOW:
                    ta_cell.fill = orange
                found_by_ta = True
                job.matched_by_ta += 1

        if not found_by_isbn and not found_by_ta:
            job.not_found += 1
        if row_rate_limited:
            job.rate_limited += 1
            # Ne pas écraser une source trouvée : ne marquer que si rien par ISBN.
            if not found_by_isbn:
                ws.cell(row=row, column=out["SOURCE_BY_ISBN"], value="RATE_LIMITED")

        job.processed += 1
        if job.processed % 10 == 0:
            job.save(update_fields=[
                "processed", "matched_by_isbn", "matched_by_ta",
                "not_found", "rate_limited",
            ])

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    job.result_file.save(f"verify-{job.pk}.xlsx", ContentFile(buffer.getvalue()), save=False)


def _parse_row_overrides(
    row,
    override_cols: dict,
    resolved_category,
    external_codes: set[str] | None = None,
    provenances: dict | None = None,
) -> tuple[dict, list[str]]:
    """FEAT-053 : extrait les overrides fiche/exemplaire d'une ligne Excel.

    Ne retient que les colonnes présentes ET non vides (cellule vide → l'info
    existante est conservée). Renvoie ``(overrides, warnings)``. AUTHOR et TAGS
    sont des listes (remplacement, pas fusion). ``resolved_category`` est la
    Category déjà résolue depuis la colonne CATEGORY (None si absente/inconnue).

    ``external_codes`` (FEAT-063) est l'ensemble des codes Ofelia externes déjà
    pris — en base au démarrage du job, puis enrichi ligne après ligne. Un code
    déjà pris n'est pas appliqué : deux exemplaires portant le même code
    rendraient tout scan ambigu.

    ``provenances`` (FEAT-064) indexe les provenances connues par code **et**
    par libellé normalisés, chargées une seule fois pour tout le fichier.
    """
    def _cell(field: str) -> str:
        idx = override_cols.get(field)
        if idx and idx <= len(row):
            return _cell_str(row[idx - 1])
        return ""

    overrides: dict = {}
    warnings: list[str] = []

    title = _cell("title")
    if title:
        overrides["title"] = title[:300]

    # CATEGORY : déjà résolue en amont (partage la colonne avec l'import de base).
    if override_cols.get("category") and resolved_category is not None:
        overrides["category"] = resolved_category

    author = _cell("author")
    if author:
        overrides["authors"] = _split_multi(author, ";")

    editor = _cell("editor")
    if editor:
        overrides["publisher"] = editor[:200]

    language = _cell("language")
    if language:
        overrides["language"] = language[:10]

    year = _cell("year")
    if year:
        try:
            overrides["publication_year"] = int(float(year))
        except (ValueError, TypeError):
            warnings.append("YEAR_INVALID")

    doc_type = _cell("type")
    if doc_type:
        resolved = _resolve_document_type(doc_type)
        if resolved:
            overrides["document_type"] = resolved
        else:
            warnings.append("TYPE_UNKNOWN")

    tags = _cell("tags")
    if tags:
        parsed = [t[:_MAX_TAG_LENGTH] for t in _split_multi(tags, ",")][:_MAX_TAGS_PER_RECORD]
        if parsed:
            overrides["tags"] = parsed

    condition = _cell("condition")
    if condition:
        resolved = _resolve_item_state(condition)
        if resolved:
            overrides["state"] = resolved
        else:
            warnings.append("CONDITION_UNKNOWN")

    # FEAT-067 : la cote appartient à la catégorie, pas à la ligne — on la pose
    # sur la Category résolue par la colonne CATEGORY. Sans catégorie résolue,
    # elle n'a pas de cible : on le signale plutôt que de la perdre en silence.
    abbr = _cell("category_abbr")
    if abbr:
        if resolved_category is None:
            warnings.append("CATEGORY_ABBR_ORPHAN")
        elif resolved_category.abbreviation != abbr[:20]:
            resolved_category.abbreviation = abbr[:20]
            resolved_category.save(update_fields=["abbreviation"])

    provenance_raw = _cell("provenance")
    if provenance_raw:
        provenance = (provenances or {}).get(_norm(provenance_raw))
        if provenance is not None:
            overrides["provenance"] = provenance
        else:
            warnings.append("PROVENANCE_UNKNOWN")

    external_code = normalize_external_code(_cell("external_code"))
    if external_code:
        if not is_valid_external_code(external_code):
            warnings.append("EXTERNAL_CODE_INVALID")
        elif external_codes is not None and external_code in external_codes:
            warnings.append("EXTERNAL_CODE_DUPLICATE")
        else:
            overrides["external_code"] = external_code
            if external_codes is not None:
                external_codes.add(external_code)

    return overrides, warnings


# Champs scalaires de la notice pilotables depuis un fichier Excel (import
# comme mise à jour). AUTHOR et TAGS sont des M2M, traités à part.
_RECORD_SCALARS = (
    "title", "publisher", "publication_year", "language", "document_type", "category",
)


def _apply_import_overrides(job, session, overrides_by_local: dict[str, dict]) -> None:
    """FEAT-053 : applique les overrides aux notices/exemplaires produits.

    Après ``finalize_scan_session``, chaque ScanItem porte dans son
    ``processing_result`` l'``record_id`` et la liste ``copies_created``. On
    écrase les champs de la notice (même préexistante) avec les valeurs Excel ;
    AUTHOR / TAGS sont remplacés. ``state`` s'applique aux exemplaires du lot.
    """
    from django.db import transaction

    from .models import Author, BibliographicRecord, Item

    if not overrides_by_local:
        return

    items = {
        si.local_id: si
        for si in ScanItem.objects.filter(
            session=session, local_id__in=list(overrides_by_local.keys())
        )
    }

    with transaction.atomic():
        for local_id, overrides in overrides_by_local.items():
            scan_item = items.get(local_id)
            if scan_item is None:
                continue
            result = scan_item.processing_result or {}
            record_id = result.get("record_id")
            copy_ids = result.get("copies_created") or []

            if record_id:
                record = BibliographicRecord.objects.filter(pk=record_id).first()
                if record:
                    changed = []
                    for field in _RECORD_SCALARS:
                        if field == "category":
                            if "category" in overrides:
                                record.category = overrides["category"]
                                changed.append("category")
                        elif field in overrides:
                            setattr(record, field, overrides[field])
                            changed.append(field)
                    if changed:
                        record.save(update_fields=changed)
                    if "authors" in overrides:
                        record.authors.clear()
                        for name in overrides["authors"]:
                            author, _ = Author.objects.get_or_create(full_name=name)
                            record.authors.add(author)
                    if "tags" in overrides:
                        record.tags.clear()
                        for name in overrides["tags"]:
                            record.tags.add(_get_or_create_tag(name))

            if "state" in overrides and copy_ids:
                Item.objects.filter(pk__in=copy_ids).update(state=overrides["state"])

            if "provenance" in overrides and copy_ids:
                Item.objects.filter(pk__in=copy_ids).update(
                    provenance=overrides["provenance"]
                )

            # FEAT-063 : le code externe désigne UN exemplaire — s'il y a
            # plusieurs copies sur la ligne, il va sur la première.
            if "external_code" in overrides and copy_ids:
                code = overrides["external_code"]
                target = copy_ids[0]
                already_taken = (
                    Item.objects.filter(external_code=code)
                    .exclude(pk=target)
                    .exists()
                )
                if not already_taken:
                    Item.objects.filter(pk=target).update(external_code=code)


def run_import_job(job: ExcelCatalogJob) -> None:
    """Mode IMPORT : crée une ScanSession virtuelle + ses ScanItem, finalise."""
    import openpyxl

    from apps.api.services import finalize_scan_session
    from .models import Location

    wb = openpyxl.load_workbook(job.uploaded_file.path, read_only=True, data_only=True)
    ws = wb.active
    headers = _header_map(ws)
    col_isbn = headers.get("isbn")
    col_loc = headers.get("location")
    col_cat = headers.get("category")
    # FEAT-053 : colonnes optionnelles d'affectation fiche/exemplaire.
    override_cols = {
        name: _resolve_column(headers, name) for name in IMPORT_OVERRIDE_COLUMNS
    }
    # local_id → dict d'overrides (uniquement les colonnes présentes ET non vides).
    overrides_by_local: dict[str, dict] = {}

    # Ré-exécution (admin) après un échec à mi-parcours : on réutilise la
    # session déjà créée → `update_or_create` sur (session, local_id) garantit
    # l'absence de doublons (risque #5 de la spec).
    session = job.scan_session
    if session is None:
        session = ScanSession.objects.create(
            label=f"Import Excel — {timezone.localtime(timezone.now()):%Y-%m-%d %H:%M}",
            created_by=job.created_by,
            state=ScanSessionState.OPEN,
        )
        job.scan_session = session
        job.save(update_fields=["scan_session"])

    known_locations = {loc.code for loc in Location.objects.all()}
    # FEAT-063 : codes externes déjà attribués (en base), enrichis au fil des
    # lignes pour attraper aussi les doublons internes au fichier.
    from .models import Item as _Item

    external_codes = set(
        _Item.objects.exclude(external_code="").values_list("external_code", flat=True)
    )
    # FEAT-064 : provenances indexées par code et par libellé (une seule requête).
    from .models import Provenance

    provenances: dict = {}
    for prov in Provenance.objects.all():
        provenances[_norm(prov.code)] = prov
        if prov.label:
            provenances.setdefault(_norm(prov.label), prov)
    total = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_isbn = _cell_str(row[col_isbn - 1]) if col_isbn and col_isbn <= len(row) else ""
        loc_code = _cell_str(row[col_loc - 1]) if col_loc and col_loc <= len(row) else ""
        cat_name = _cell_str(row[col_cat - 1]) if col_cat and col_cat <= len(row) else ""
        if not raw_isbn:
            # BUG-025 : une ligne **remplie mais sans ISBN** était sautée en
            # silence — absente des compteurs comme du rapport, donc « disparue »
            # (105 lignes dans le fichier → 104 notices, sans une seule erreur).
            # L'import est indexé par ISBN, on ne peut pas la créer ici : on la
            # signale pour qu'elle soit cataloguée à la main. Les lignes
            # entièrement vides (openpyxl en compte souvent après les données)
            # restent ignorées sans bruit.
            if any(_cell_str(cell) for cell in row):
                total += 1
                job.processed += 1
                job.errors += 1
                job.report.append({
                    "row": row_idx,
                    "isbn": "",
                    "warning": "ISBN_MISSING",
                    "label": _row_label(row, headers),
                })
            continue
        total += 1
        isbn = normalize_isbn(raw_isbn)
        if len(isbn) not in (10, 13):
            job.errors += 1
            job.report.append({
                "row": row_idx,
                "isbn": raw_isbn,
                "warning": "ISBN_INVALID",
                "label": _row_label(row, headers),
            })
            job.processed += 1
            continue

        warnings = []
        if loc_code and loc_code not in known_locations:
            warnings.append("LOCATION_UNKNOWN")
            loc_code = ""
        category = None
        if cat_name:
            category = _resolve_category(cat_name)
            if category is None:
                warnings.append("CATEGORY_UNKNOWN")

        # FEAT-053 : overrides fiche/exemplaire (colonnes présentes + non vides).
        local_id = f"excel-{job.pk}-{row_idx}"
        overrides, ov_warnings = _parse_row_overrides(
            row, override_cols, category, external_codes, provenances
        )
        warnings.extend(ov_warnings)
        if overrides:
            overrides_by_local[local_id] = overrides

        ScanItem.objects.update_or_create(
            session=session,
            local_id=local_id,
            defaults={
                "scan_kind": ScanKind.ISBN if len(isbn) == 10 else ScanKind.EAN13,
                "scanned_value": isbn,
                # FEAT-053 : titre du fichier (sinon placeholder posé par
                # _create_record, remplaçable ensuite par l'enrichissement).
                "metadata_title": overrides.get("title", ""),
                "category": category,
                "location_code": loc_code,
                "copy_count": 1,
                "scanned_at": timezone.now(),
            },
        )
        if warnings:
            job.report.append({"row": row_idx, "isbn": isbn, "warning": ", ".join(warnings)})
        job.processed += 1
    wb.close()

    job.total = total
    job.save(update_fields=["total", "processed", "errors", "report"])

    summary = finalize_scan_session(session)
    # FEAT-053 : après matérialisation, appliquer les overrides aux notices/
    # exemplaires produits (écrase l'existant le cas échéant).
    _apply_import_overrides(job, session, overrides_by_local)
    job.matched_by_isbn = summary.get("records_matched", 0)
    job.not_found = summary.get("records_created", 0)
    job.save(update_fields=["matched_by_isbn", "not_found"])


# ─── FEAT-079 : mise à jour d'exemplaires existants ──────────────────────


def _cell_at(row, col: int | None) -> str:
    """Valeur texte de la colonne `col` (1-based) de `row`, "" si absente."""
    if col and col <= len(row):
        return _cell_str(row[col - 1])
    return ""


def _find_item_by_ofelia_code(raw: str):
    """Exemplaire désigné par un code Ofelia, ou None.

    Deux écritures circulent pour le même exemplaire : l'EAN13 « 290… » que
    porte le code-barres de l'étiquette, et le code interne « OFL-… » imprimé
    en clair juste à côté. L'export écrit les deux ; un fichier repris à la
    main peut ne contenir que l'un ou l'autre, on essaie donc les deux.
    """
    from .models import Item

    value = (raw or "").strip()
    if not value:
        return None
    code = normalize_external_code(value)
    return (
        Item.objects.filter(ean13=code).first()
        or Item.objects.filter(internal_id__iexact=value).first()
    )


def _apply_item_update(item, overrides, location_value, isbn_value, locations):
    """Applique une ligne de mise à jour à `item`. Renvoie ``(changed, warnings)``.

    Une cellule vide ne touche à rien (même règle qu'à l'import) : le fichier
    exporté peut être renvoyé avec deux colonnes corrigées sans que les autres
    soient réécrites. Chaque champ est comparé avant écriture, pour que
    « lignes mises à jour » compte les changements réels et pas les lignes lues.
    """
    from .models import Author, BibliographicRecord, Item

    warnings: list[str] = []
    record = item.record
    record_changed: list[str] = []
    item_changed: list[str] = []

    for field in _RECORD_SCALARS:
        if field not in overrides:
            continue
        if field == "category":
            wanted = overrides["category"]
            if record.category_id != (wanted.pk if wanted else None):
                record.category = wanted
                record_changed.append("category")
        elif getattr(record, field) != overrides[field]:
            setattr(record, field, overrides[field])
            record_changed.append(field)

    # ISBN : en import c'est la clé de la ligne, en mise à jour c'est un champ
    # comme un autre — une coquille d'ISBN se corrige dans le fichier exporté.
    if isbn_value:
        isbn = normalize_isbn(isbn_value)
        if len(isbn) not in (10, 13):
            warnings.append("ISBN_INVALID")
        else:
            field = "isbn_13" if len(isbn) == 13 else "isbn_10"
            if (getattr(record, field) or "") != isbn:
                # L'unicité d'isbn_13 est garantie en base : sans ce garde-fou,
                # la ligne ferait planter tout le lot au lieu d'être signalée.
                taken = (
                    field == "isbn_13"
                    and BibliographicRecord.objects.filter(isbn_13=isbn)
                    .exclude(pk=record.pk)
                    .exists()
                )
                if taken:
                    warnings.append("ISBN_CONFLICT")
                else:
                    setattr(record, field, isbn)
                    record_changed.append(field)

    m2m_changed = False
    if "authors" in overrides:
        wanted = overrides["authors"]
        if sorted(a.full_name for a in record.authors.all()) != sorted(wanted):
            record.authors.set(
                [Author.objects.get_or_create(full_name=name)[0] for name in wanted]
            )
            m2m_changed = True
    if "tags" in overrides:
        wanted = overrides["tags"]
        if sorted(t.name for t in record.tags.all()) != sorted(wanted):
            record.tags.set([_get_or_create_tag(name) for name in wanted])
            m2m_changed = True

    if "state" in overrides and item.state != overrides["state"]:
        item.state = overrides["state"]
        item_changed.append("state")

    if "provenance" in overrides and item.provenance_id != overrides["provenance"].pk:
        item.provenance = overrides["provenance"]
        item_changed.append("provenance")

    code = overrides.get("external_code")
    if code and item.external_code != code:
        # Un code externe désigne un exemplaire et un seul (contrainte partielle
        # d'unicité) : si deux lignes du fichier se disputent le même code, la
        # première l'emporte et la seconde est signalée.
        if Item.objects.filter(external_code=code).exclude(pk=item.pk).exists():
            warnings.append("EXTERNAL_CODE_DUPLICATE")
        else:
            item.external_code = code
            item_changed.append("external_code")

    if location_value:
        location = locations.get(_norm(location_value))
        if location is None:
            warnings.append("LOCATION_UNKNOWN")
        elif item.location_id != location.pk:
            item.location = location
            item_changed.append("location")

    if record_changed:
        record.save(update_fields=record_changed + ["updated_at"])
    if item_changed:
        item.save(update_fields=item_changed + ["updated_at"])

    return bool(record_changed or item_changed or m2m_changed), warnings


def run_update_job(job: ExcelCatalogJob) -> None:
    """Mode UPDATE (FEAT-079) : met à jour des exemplaires **existants**.

    Aucune création, jamais : une ligne dont l'exemplaire est introuvable est
    comptée en erreur et listée dans le rapport, pas transformée en nouveau
    livre. C'est toute la différence avec l'import, et c'est ce qui permet de
    renvoyer un fichier exporté sans risquer de dupliquer la bibliothèque.

    Clé de la ligne : code Ofelia (EAN13 « 290… » ou code interne « OFL-… »)
    et/ou code externe. Si les deux sont renseignés, **le code Ofelia gagne** —
    c'est lui qui désigne l'exemplaire, et le code externe de la ligne lui est
    alors appliqué (c'est la façon d'attribuer un code externe en masse).
    """
    import openpyxl
    from django.db import transaction

    from .models import Item, Location, Provenance

    wb = openpyxl.load_workbook(job.uploaded_file.path, read_only=True, data_only=True)
    ws = wb.active
    headers = _header_map(ws)
    key_cols = {name: _resolve_column(headers, name) for name in UPDATE_KEY_COLUMNS}
    override_cols = {
        name: _resolve_column(headers, name) for name in UPDATE_OVERRIDE_COLUMNS
    }

    # Référentiels chargés une fois pour tout le fichier (une bibliothèque a
    # quelques dizaines d'emplacements et de provenances, pas des milliers).
    locations = {_norm(loc.code): loc for loc in Location.objects.all()}
    provenances: dict = {}
    for prov in Provenance.objects.all():
        provenances[_norm(prov.code)] = prov
        if prov.label:
            provenances.setdefault(_norm(prov.label), prov)

    total = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # openpyxl compte souvent des lignes vides après les données.
        if not any(_cell_str(cell) for cell in row):
            continue
        total += 1

        ofelia_raw = _cell_at(row, key_cols.get("ofelia_code")) or _cell_at(
            row, key_cols.get("internal_id")
        )
        external_raw = normalize_external_code(
            _cell_at(row, key_cols.get("external_code"))
        )

        item = None
        failure = ""
        failure_code = ""
        if ofelia_raw:
            item = _find_item_by_ofelia_code(ofelia_raw)
            if item is None:
                # On ne retombe pas sur le code externe : un code Ofelia qui ne
                # correspond à rien veut dire que la ligne désigne mal son
                # exemplaire. Mieux vaut la signaler que modifier au jugé.
                failure, failure_code = "OFELIA_CODE_UNKNOWN", ofelia_raw
        elif external_raw:
            item = Item.objects.filter(external_code=external_raw).first()
            if item is None:
                failure, failure_code = "EXTERNAL_CODE_UNKNOWN", external_raw
        else:
            failure = "NO_KEY"

        if item is None:
            job.errors += 1
            job.processed += 1
            job.report.append({
                "row": row_idx,
                "code": failure_code,
                "warning": failure,
                "label": _row_label(row, headers),
            })
            continue

        warnings: list[str] = []
        cat_name = _cell_at(row, override_cols.get("category"))
        category = None
        if cat_name:
            category = _resolve_category(cat_name)
            if category is None:
                warnings.append("CATEGORY_UNKNOWN")

        # `_parse_row_overrides` pose la cote directement sur la catégorie : on
        # relève sa valeur avant/après, sinon un fichier qui ne corrige que des
        # cotes s'afficherait comme « 0 ligne mise à jour ».
        abbr_before = category.abbreviation if category else None
        # `external_codes=None` : l'unicité du code externe est vérifiée plus
        # bas en excluant l'exemplaire lui-même, sinon une ligne qui renvoie le
        # code déjà porté par son propre exemplaire passerait pour un doublon.
        overrides, ov_warnings = _parse_row_overrides(
            row, override_cols, category, None, provenances
        )
        warnings.extend(ov_warnings)
        abbr_changed = category is not None and category.abbreviation != abbr_before

        try:
            with transaction.atomic():
                changed, apply_warnings = _apply_item_update(
                    item,
                    overrides,
                    _cell_at(row, override_cols.get("location")),
                    _cell_at(row, override_cols.get("isbn")),
                    locations,
                )
        except Exception as exc:  # filet : une ligne qui casse n'annule pas le lot
            logger.exception("UPDATE ligne %s (%s) KO", row_idx, item.internal_id)
            job.errors += 1
            job.processed += 1
            job.report.append({
                "row": row_idx,
                "code": item.internal_id,
                "warning": "ROW_ERROR",
                "label": str(exc)[:200],
            })
            continue

        warnings.extend(apply_warnings)
        if changed or abbr_changed:
            job.updated += 1
        else:
            job.unchanged += 1
        if warnings:
            job.report.append({
                "row": row_idx,
                "code": item.internal_id,
                "warning": ", ".join(warnings),
                "label": _row_label(row, headers),
            })
        job.processed += 1
        if job.processed % 10 == 0:
            job.save(update_fields=[
                "processed", "updated", "unchanged", "errors", "report",
            ])

    wb.close()
    job.total = total
    job.save(update_fields=[
        "total", "processed", "updated", "unchanged", "errors", "report",
    ])


def run_excel_catalog_job(job_id: int) -> None:
    """Point d'entrée django-q2 : dispatch VERIFY / IMPORT.

    Idempotente : si le job n'est plus PENDING (worker concurrent via le retry
    django-q2), on s'arrête immédiatement.
    """
    job = ExcelCatalogJob.objects.get(pk=job_id)
    if job.state != ExcelJobState.PENDING:
        logger.info("ExcelCatalogJob #%s déjà %s, abandon (re-entry).", job_id, job.state)
        return
    try:
        job.state = ExcelJobState.RUNNING
        job.save(update_fields=["state"])
        if job.mode == ExcelJobMode.VERIFY:
            run_verify_job(job)
        elif job.mode == ExcelJobMode.UPDATE:
            run_update_job(job)
        else:
            run_import_job(job)
        job.state = ExcelJobState.FINISHED
        job.finished_at = timezone.now()
        job.save()
    except Exception as exc:
        logger.exception("ExcelCatalogJob %s failed", job_id)
        job.state = ExcelJobState.FAILED
        job.report.append({"global_error": str(exc)})
        job.finished_at = timezone.now()
        job.save()
