"""FEAT-050 — catalogage Excel (vérification + import)."""
import io

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile

from apps.accounts.models import Role, User
from apps.catalog import excel_catalog
from apps.catalog.excel_catalog import (
    run_import_job,
    run_verify_job,
    validate_xlsx,
)
from apps.catalog.models import (
    Author,
    BibliographicRecord,
    Category,
    DocumentType,
    ExcelCatalogJob,
    ExcelJobMode,
    Item,
    ItemState,
    Location,
    ScanItem,
    ScanSession,
    Tag,
)

pytestmark = pytest.mark.django_db

VALID_ISBN = "9782070368228"


def _xlsx(headers, rows, name="inventaire.xlsx"):
    """Construit un SimpleUploadedFile .xlsx en mémoire."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return SimpleUploadedFile(
        name,
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pytest.fixture(autouse=True)
def _tmp_media(settings, tmp_path):
    """Isole les uploads/result_file dans un dossier temporaire."""
    settings.MEDIA_ROOT = str(tmp_path)


@pytest.fixture
def user():
    return User.objects.create_user(username="lib", password="pw", role=Role.LIBRARIAN)


# ── Validation ─────────────────────────────────────────────────────────────


def test_validate_xlsx_rejects_xls():
    f = SimpleUploadedFile("old.xls", b"garbage", content_type="application/vnd.ms-excel")
    errors = validate_xlsx(f, ExcelJobMode.VERIFY)
    assert errors and "xlsx" in errors[0].lower()


def test_validate_xlsx_rejects_oversized(monkeypatch):
    monkeypatch.setattr(excel_catalog, "MAX_FILE_BYTES", 10)
    f = _xlsx(["ID", "TITLE", "AUTHOR", "ISBN"], [["1", "T", "A", VALID_ISBN]])
    errors = validate_xlsx(f, ExcelJobMode.VERIFY)
    assert any("Mo" in e for e in errors)


def test_validate_xlsx_requires_columns_verify():
    f = _xlsx(["ID", "AUTHOR", "ISBN"], [["1", "A", VALID_ISBN]])  # TITLE manquant
    errors = validate_xlsx(f, ExcelJobMode.VERIFY)
    assert any("TITLE" in e for e in errors)


def test_validate_xlsx_requires_columns_import():
    f = _xlsx(["ID", "TITLE"], [["1", "T"]])  # ISBN manquant
    errors = validate_xlsx(f, ExcelJobMode.IMPORT)
    assert any("ISBN" in e for e in errors)


def test_validate_xlsx_accents_tolerated():
    # En-têtes accentués / casse mixte → normalisés.
    f = _xlsx(["Id", "Titre", "Author", "Isbn"], [["1", "T", "A", VALID_ISBN]])
    errors = validate_xlsx(f, ExcelJobMode.IMPORT)
    assert errors == []  # ISBN présent (insensible casse)


# ── Mode VERIFY ────────────────────────────────────────────────────────────


def _make_verify_job(user, headers, rows):
    f = _xlsx(headers, rows)
    return ExcelCatalogJob.objects.create(
        mode=ExcelJobMode.VERIFY, uploaded_file=f, created_by=user
    )


def _read_result(job):
    wb = openpyxl.load_workbook(job.result_file.path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    data = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    return headers, data


def test_verify_job_pass1_only(user, monkeypatch):
    monkeypatch.setattr(
        excel_catalog,
        "_pass1_by_isbn",
        lambda isbn: ({"title": "Vrai Titre", "authors_text": "Vrai Auteur", "source": "bnf"}, False),
    )
    monkeypatch.setattr(excel_catalog, "_search_all", lambda title, author: ([], False))
    job = _make_verify_job(user, ["ID", "TITLE", "AUTHOR", "ISBN"],
                           [["1", "Saisi", "Saisi", VALID_ISBN]])
    run_verify_job(job)
    assert job.matched_by_isbn == 1
    assert job.not_found == 0
    headers, data = _read_result(job)
    assert "TITLE_FOUND_BY_ISBN" in headers
    col = headers.index("TITLE_FOUND_BY_ISBN")
    assert data[0][col] == "Vrai Titre"


def test_verify_job_pass2_runs_even_with_isbn(user, monkeypatch):
    """La passe 2 (titre+auteur) tourne même quand l'ISBN a résolu la ligne :
    permet de recouper un ISBN saisi à la main."""
    monkeypatch.setattr(
        excel_catalog,
        "_pass1_by_isbn",
        lambda isbn: ({"title": "Le Petit Prince", "authors_text": "Saint-Exupéry", "source": "bnf"}, False),
    )
    monkeypatch.setattr(
        excel_catalog,
        "_search_all",
        lambda title, author: ([
            {"title": "Le Petit Prince", "authors_text": "Antoine de Saint-Exupéry",
             "isbn_13": "9782070408504", "isbn_10": ""},
        ], False),
    )
    # ISBN saisi (VALID_ISBN) ≠ ISBN trouvé par titre+auteur → recoupement.
    job = _make_verify_job(user, ["ID", "TITLE", "AUTHOR", "ISBN"],
                           [["1", "Le Petit Prince", "Saint-Exupéry", VALID_ISBN]])
    run_verify_job(job)
    assert job.matched_by_isbn == 1
    assert job.matched_by_ta == 1  # passe 2 a aussi tourné
    headers, data = _read_result(job)
    col = headers.index("ISBN_FOUND_BY_TA")
    assert data[0][col] == "9782070408504"


def test_verify_job_pass2_fuzzy_match(user, monkeypatch):
    monkeypatch.setattr(excel_catalog, "_pass1_by_isbn", lambda isbn: (None, False))
    monkeypatch.setattr(
        excel_catalog,
        "_search_all",
        lambda title, author: ([
            {"title": "Le Petit Prince", "authors_text": "Antoine de Saint-Exupéry",
             "isbn_13": VALID_ISBN, "isbn_10": ""},
        ], False),
    )
    # Ligne sans ISBN, titre avec faute de frappe → doit matcher (score > 60).
    job = _make_verify_job(user, ["ID", "TITLE", "AUTHOR", "ISBN"],
                           [["1", "Le Petit Pince", "Saint Exupery", ""]])
    run_verify_job(job)
    assert job.matched_by_ta == 1
    assert job.not_found == 0
    headers, data = _read_result(job)
    col = headers.index("ISBN_FOUND_BY_TA")
    assert data[0][col] == VALID_ISBN


def test_verify_job_pass2_low_score_skipped(user, monkeypatch):
    monkeypatch.setattr(excel_catalog, "_pass1_by_isbn", lambda isbn: (None, False))
    monkeypatch.setattr(
        excel_catalog,
        "_search_all",
        lambda title, author: ([
            {"title": "Un livre totalement différent zzz", "authors_text": "Quelqu'un d'autre",
             "isbn_13": "9780000000000", "isbn_10": ""},
        ], False),
    )
    job = _make_verify_job(user, ["ID", "TITLE", "AUTHOR", "ISBN"],
                           [["1", "Mon titre unique abc", "Moi", ""]])
    run_verify_job(job)
    assert job.matched_by_ta == 0
    assert job.not_found == 1
    headers, data = _read_result(job)
    col = headers.index("ISBN_FOUND_BY_TA")
    assert not data[0][col]  # rien écrit


# ── Mode IMPORT ────────────────────────────────────────────────────────────


def _make_import_job(user, headers, rows):
    f = _xlsx(headers, rows)
    return ExcelCatalogJob.objects.create(
        mode=ExcelJobMode.IMPORT, uploaded_file=f, created_by=user
    )


def test_import_job_creates_scan_session_and_items(user):
    job = _make_import_job(user, ["ISBN"], [[VALID_ISBN], ["9782266283434"]])
    run_import_job(job)
    job.refresh_from_db()
    assert job.scan_session is not None
    assert ScanItem.objects.filter(session=job.scan_session).count() == 2
    # finalize_scan_session a matérialisé les notices + exemplaires.
    assert Item.objects.count() == 2


def test_import_job_resolves_location_and_category(user):
    Location.objects.create(code="A1")
    cat = Category.objects.create(code="ROM", name="Roman")
    job = _make_import_job(user, ["ISBN", "LOCATION", "CATEGORY"],
                           [[VALID_ISBN, "A1", "Roman"]])
    run_import_job(job)
    item = ScanItem.objects.get(session=job.scan_session)
    assert item.location_code == "A1"
    assert item.category_id == cat.pk


def test_import_job_skips_invalid_isbn(user):
    job = _make_import_job(user, ["ISBN"], [["not-an-isbn"], [VALID_ISBN]])
    run_import_job(job)
    job.refresh_from_db()
    assert job.errors == 1
    assert ScanItem.objects.filter(session=job.scan_session).count() == 1
    assert any(e.get("warning") == "ISBN_INVALID" for e in job.report)


def test_import_job_unknown_location_warns(user):
    job = _make_import_job(user, ["ISBN", "LOCATION"], [[VALID_ISBN, "ZZ9"]])
    run_import_job(job)
    job.refresh_from_db()
    item = ScanItem.objects.get(session=job.scan_session)
    assert item.location_code == ""  # emplacement inconnu ignoré
    assert any("LOCATION_UNKNOWN" in e.get("warning", "") for e in job.report)


def test_index_page_renders(client, user):
    client.force_login(user)
    from django.urls import reverse

    resp = client.get(reverse("catalog:excel_catalog_index"))
    assert resp.status_code == 200
    assert b"Catalogage Excel" in resp.content


def test_detail_page_renders(client, user):
    client.force_login(user)
    from django.urls import reverse

    job = ExcelCatalogJob.objects.create(mode=ExcelJobMode.VERIFY, created_by=user)
    resp = client.get(reverse("catalog:excel_catalog_detail", args=[job.pk]))
    assert resp.status_code == 200


def test_detail_other_user_redirects(client):
    other = User.objects.create_user(username="other", password="pw", role=Role.LIBRARIAN)
    owner = User.objects.create_user(username="owner", password="pw", role=Role.LIBRARIAN)
    job = ExcelCatalogJob.objects.create(mode=ExcelJobMode.VERIFY, created_by=owner)
    client.force_login(other)
    from django.urls import reverse

    resp = client.get(reverse("catalog:excel_catalog_detail", args=[job.pk]))
    assert resp.status_code == 302  # pas le propriétaire → renvoyé à l'index


# ── FEAT-053 : colonnes métadonnées fiche / exemplaire ──────────────────────


def test_import_overrides_new_record_all_fields(user):
    cat = Category.objects.create(code="ROM", name="Roman")
    headers = ["ISBN", "TITLE", "AUTHOR", "CATEGORY", "TYPE", "EDITOR", "YEAR", "LANGUAGE", "TAGS", "CONDITION"]
    rows = [[VALID_ISBN, "Les Misérables", "Victor Hugo; Alexandre Dumas", "Roman", "BD / manga",
             "Gallimard", "1998", "en", "aventure, classique", "Usé"]]
    job = _make_import_job(user, headers, rows)
    run_import_job(job)
    rec = BibliographicRecord.objects.get(isbn_13=VALID_ISBN)
    assert rec.title == "Les Misérables"  # pas de placeholder ISBN:…
    assert rec.category_id == cat.pk
    assert rec.document_type == DocumentType.COMIC
    assert rec.publisher == "Gallimard"
    assert rec.publication_year == 1998
    assert rec.language == "en"
    assert set(rec.authors.values_list("full_name", flat=True)) == {"Victor Hugo", "Alexandre Dumas"}
    assert set(rec.tags.values_list("name", flat=True)) == {"aventure", "classique"}
    item = Item.objects.get(record=rec)
    assert item.state == ItemState.WORN


def test_import_overrides_existing_record(user):
    """Colonne présente + non vide → écrase la fiche existante (même ISBN)."""
    old_author = Author.objects.create(full_name="Ancien Auteur")
    old_tag = Tag.objects.create(name="ancien")
    rec = BibliographicRecord.objects.create(
        title="Titre existant", isbn_13=VALID_ISBN, publisher="Vieil éditeur",
        publication_year=1900, language="es",
    )
    rec.authors.add(old_author)
    rec.tags.add(old_tag)
    headers = ["ISBN", "AUTHOR", "EDITOR", "YEAR", "TAGS"]
    rows = [[VALID_ISBN, "Nouvel Auteur", "Nouvel éditeur", "2020", "neuf, frais"]]
    job = _make_import_job(user, headers, rows)
    run_import_job(job)
    rec.refresh_from_db()
    assert rec.publisher == "Nouvel éditeur"
    assert rec.publication_year == 2020
    # Titre non touché (pas de colonne TITLE) — l'existant reste.
    assert rec.title == "Titre existant"
    # AUTHOR / TAGS remplacés, pas fusionnés.
    assert set(rec.authors.values_list("full_name", flat=True)) == {"Nouvel Auteur"}
    assert set(rec.tags.values_list("name", flat=True)) == {"neuf", "frais"}


def test_import_title_overwrites_existing(user):
    """Colonne TITLE remplie → écrase le titre de la fiche existante."""
    rec = BibliographicRecord.objects.create(title="Ancien titre", isbn_13=VALID_ISBN)
    job = _make_import_job(user, ["ISBN", "TITLE"], [[VALID_ISBN, "Titre corrigé"]])
    run_import_job(job)
    rec.refresh_from_db()
    assert rec.title == "Titre corrigé"


def test_import_no_title_column_uses_placeholder(user):
    """Sans colonne TITLE, une notice neuve garde le placeholder ISBN:… (FEAT-050)."""
    job = _make_import_job(user, ["ISBN"], [[VALID_ISBN]])
    run_import_job(job)
    rec = BibliographicRecord.objects.get(isbn_13=VALID_ISBN)
    assert rec.title.startswith("ISBN:")


def test_import_empty_cell_keeps_existing(user):
    """Colonne présente mais cellule vide → l'info existante est conservée."""
    rec = BibliographicRecord.objects.create(
        title="T", isbn_13=VALID_ISBN, publisher="Éditeur d'origine", language="mg",
    )
    headers = ["ISBN", "EDITOR", "LANGUAGE"]
    rows = [[VALID_ISBN, "", ""]]  # cellules vides
    job = _make_import_job(user, headers, rows)
    run_import_job(job)
    rec.refresh_from_db()
    assert rec.publisher == "Éditeur d'origine"
    assert rec.language == "mg"


def test_import_unknown_type_and_condition_warn(user):
    headers = ["ISBN", "TYPE", "CONDITION"]
    rows = [[VALID_ISBN, "hologramme", "parfait"]]
    job = _make_import_job(user, headers, rows)
    run_import_job(job)
    job.refresh_from_db()
    warns = " ".join(e.get("warning", "") for e in job.report)
    assert "TYPE_UNKNOWN" in warns
    assert "CONDITION_UNKNOWN" in warns
    rec = BibliographicRecord.objects.get(isbn_13=VALID_ISBN)
    assert rec.document_type == DocumentType.BOOK  # défaut inchangé
    assert Item.objects.get(record=rec).state == ItemState.GOOD  # défaut inchangé


def test_import_invalid_year_warns(user):
    headers = ["ISBN", "YEAR"]
    rows = [[VALID_ISBN, "mille-neuf-cent"]]
    job = _make_import_job(user, headers, rows)
    run_import_job(job)
    job.refresh_from_db()
    assert any("YEAR_INVALID" in e.get("warning", "") for e in job.report)
    rec = BibliographicRecord.objects.get(isbn_13=VALID_ISBN)
    assert rec.publication_year is None


def test_import_type_by_code(user):
    headers = ["ISBN", "TYPE"]
    rows = [[VALID_ISBN, "audio_cd"]]
    job = _make_import_job(user, headers, rows)
    run_import_job(job)
    rec = BibliographicRecord.objects.get(isbn_13=VALID_ISBN)
    assert rec.document_type == DocumentType.AUDIO_CD


def test_import_job_idempotent_on_local_id(user):
    job = _make_import_job(user, ["ISBN"], [[VALID_ISBN]])
    run_import_job(job)
    job.refresh_from_db()
    count_after_first = ScanItem.objects.filter(session=job.scan_session).count()
    # Ré-exécution sur la même session (même job) → pas de doublon.
    run_import_job(job)
    assert ScanItem.objects.filter(session=job.scan_session).count() == count_after_first


# ── BUG-025 : ligne sans ISBN ──────────────────────────────────────────────


def test_import_job_reports_row_without_isbn(user):
    """Une ligne remplie mais sans ISBN doit être **signalée**, pas escamotée.

    Cas réel : 105 lignes dans le fichier, 104 notices créées, 0 erreur — la
    ligne sans ISBN disparaissait sans laisser de trace ni dans les compteurs
    ni dans le rapport.
    """
    job = _make_import_job(
        user,
        ["ISBN", "AUTHOR", "TITLE"],
        [
            [VALID_ISBN, "Auteur A", "Titre A"],
            ["", "Ruiz, Miguel", "L'art de vivre et de mourir"],
            ["9782266283434", "Auteur B", "Titre B"],
        ],
    )
    run_import_job(job)
    job.refresh_from_db()

    assert job.total == 3           # la ligne sans ISBN est comptée
    assert job.processed == 3
    assert job.errors == 1
    entries = [e for e in job.report if e.get("warning") == "ISBN_MISSING"]
    assert len(entries) == 1
    assert entries[0]["row"] == 3
    # Le rapport doit permettre d'identifier le livre à cataloguer à la main.
    assert "Ruiz, Miguel" in entries[0]["label"]
    # Les autres lignes sont importées normalement.
    assert ScanItem.objects.filter(session=job.scan_session).count() == 2


def test_import_job_ignores_fully_empty_rows(user):
    """Les lignes entièrement vides (openpyxl en compte après les données)
    restent ignorées sans bruit."""
    job = _make_import_job(user, ["ISBN", "AUTHOR"], [[VALID_ISBN, "A"], ["", ""]])
    run_import_job(job)
    job.refresh_from_db()
    assert job.total == 1
    assert job.errors == 0
    assert job.report == []
