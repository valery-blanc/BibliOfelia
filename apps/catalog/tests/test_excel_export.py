"""FEAT-078 — export Excel du catalogue (une ligne par exemplaire)."""
import io

import openpyxl
import pytest
from django.urls import reverse

from apps.accounts.models import Role, User
from apps.catalog.excel_export import EXPORT_COLUMNS, build_catalog_workbook
from apps.catalog.models import (
    Author,
    BibliographicRecord,
    Category,
    DocumentType,
    Item,
    ItemState,
    Location,
    Provenance,
    Tag,
)

pytestmark = pytest.mark.django_db


def _read(content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    return headers, rows


@pytest.fixture
def catalogue():
    """Une notice complète, deux exemplaires — l'un rangé, l'autre non."""
    category = Category.objects.create(code="RO", name="Romans", abbreviation="RO FI")
    location = Location.objects.create(code="A1")
    provenance = Provenance.objects.create(code="DON2024", label="Don Dupont")
    record = BibliographicRecord.objects.create(
        title="Le Petit Prince",
        isbn_13="9782070612758",
        publisher="Gallimard",
        publication_year=1943,
        language="fr",
        category=category,
        document_type=DocumentType.BOOK,
    )
    record.authors.add(Author.objects.create(full_name="Antoine de Saint-Exupéry"))
    record.tags.add(Tag.objects.create(name="classique"))
    first = Item.objects.create(
        record=record,
        location=location,
        provenance=provenance,
        state=ItemState.GOOD,
        external_code="BCF13298781X",
    )
    second = Item.objects.create(record=record, state=ItemState.WORN)
    return record, first, second


def test_export_headers_are_the_import_columns():
    headers, rows = _read(build_catalog_workbook())
    assert headers == EXPORT_COLUMNS
    assert rows == []


def test_export_one_row_per_item(catalogue):
    _record, first, second = catalogue
    _headers, rows = _read(build_catalog_workbook())
    assert len(rows) == 2
    codes = {row[EXPORT_COLUMNS.index("INTERNAL_ID")] for row in rows}
    assert codes == {first.internal_id, second.internal_id}


def test_export_row_carries_every_field(catalogue):
    _record, first, _second = catalogue
    _headers, rows = _read(build_catalog_workbook())
    row = next(r for r in rows if r[EXPORT_COLUMNS.index("INTERNAL_ID")] == first.internal_id)
    values = dict(zip(EXPORT_COLUMNS, row))
    assert values["OFELIA_CODE"] == first.ean13
    assert values["EXTERNAL_CODE"] == "BCF13298781X"
    assert values["ISBN"] == "9782070612758"
    assert values["TITLE"] == "Le Petit Prince"
    assert values["AUTHOR"] == "Antoine de Saint-Exupéry"
    assert values["CATEGORY"] == "Romans"
    assert values["CATEGORY_ABBR"] == "RO FI"
    assert values["TYPE"] == "Livre"
    assert values["EDITOR"] == "Gallimard"
    assert values["YEAR"] == 1943
    assert values["LANGUAGE"] == "fr"
    assert values["TAGS"] == "classique"
    assert values["CONDITION"] == "Bon"
    assert values["PROVENANCE"] == "DON2024"
    assert values["LOCATION"] == "A1"


def test_export_item_without_optional_relations(catalogue):
    """Un exemplaire sans emplacement ni provenance sort avec des cellules vides,
    pas avec « None » : la cellule doit rester vide au retour en mise à jour."""
    _record, _first, second = catalogue
    _headers, rows = _read(build_catalog_workbook())
    row = next(r for r in rows if r[EXPORT_COLUMNS.index("INTERNAL_ID")] == second.internal_id)
    values = dict(zip(EXPORT_COLUMNS, row))
    assert values["LOCATION"] in (None, "")
    assert values["PROVENANCE"] in (None, "")
    assert values["EXTERNAL_CODE"] in (None, "")


def test_export_multi_author_and_tags_separators(catalogue):
    """AUTHOR se relit avec « ; » et TAGS avec « , » (règles de l'import)."""
    record, _first, _second = catalogue
    record.authors.add(Author.objects.create(full_name="Deuxième Auteur"))
    record.tags.add(Tag.objects.create(name="jeunesse"))
    _headers, rows = _read(build_catalog_workbook())
    values = dict(zip(EXPORT_COLUMNS, rows[0]))
    assert "; " in values["AUTHOR"]
    assert ", " in values["TAGS"]


def test_export_view_returns_xlsx_attachment(client, catalogue):
    User.objects.create_user(username="lib", password="pw", role=Role.LIBRARIAN)
    client.login(username="lib", password="pw")
    resp = client.get(reverse("catalog:excel_catalog_export"))
    assert resp.status_code == 200
    assert "spreadsheetml" in resp["Content-Type"]
    assert "attachment" in resp["Content-Disposition"]
    assert ".xlsx" in resp["Content-Disposition"]
    headers, rows = _read(resp.content)
    assert headers == EXPORT_COLUMNS
    assert len(rows) == 2


def test_export_view_refused_to_readonly(client, catalogue):
    User.objects.create_user(username="ro", password="pw", role=Role.READONLY)
    client.login(username="ro", password="pw")
    resp = client.get(reverse("catalog:excel_catalog_export"))
    assert resp.status_code == 403
